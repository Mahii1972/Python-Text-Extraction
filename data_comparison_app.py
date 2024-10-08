import streamlit as st
import pandas as pd
import re
import pdfplumber
from bs4 import BeautifulSoup
import msoffcrypto
import io
from fuzzywuzzy import fuzz
import openpyxl

# Function to clean phone numbers
def clean_phone_number(phone):
    return re.sub(r'\D', '', str(phone).split('.')[0])

# Function to extract text from PDF
def extract_text_from_pdf(uploaded_file):
    text = ''
    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            text += page.extract_text()
    return text

# Function to extract text from TXT
def extract_text_from_txt(uploaded_file):
    text = uploaded_file.read().decode("utf-8") 
    return text

# Function to extract text from HTML
def extract_text_from_html(uploaded_file):
    soup = BeautifulSoup(uploaded_file, 'html.parser')
    text = soup.get_text()
    return text

# Modified function to extract text and cell locations from Excel
def extract_from_excel(uploaded_file, password=None):
    try:
        if password:
            decrypted = io.BytesIO()
            file = msoffcrypto.OfficeFile(uploaded_file)
            file.load_key(password=password)
            file.decrypt(decrypted)
            workbook = openpyxl.load_workbook(decrypted, data_only=True)
        else:
            workbook = openpyxl.load_workbook(uploaded_file, data_only=True)
        
        data = []
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value:
                        data.append((str(cell.value), f"{sheet}!{cell.coordinate}", cell.value))
        return data
    except msoffcrypto.exceptions.InvalidKeyError:
        return "encrypted"
    except Exception as e:
        if "Workbook is encrypted" in str(e) or "Can't find workbook in OLE2 compound document" in str(e):
            return "encrypted"
        else:
            st.error(f"Failed to read Excel file: {e}")
            return None

# Updated function to compare data with word-by-word matching for Excel
def compare_data(base_df, compare_data, selected_columns, match_ratio, is_excel=False):
    results = []
    total_comparisons = len(base_df) * len(selected_columns)
    progress_bar = st.progress(0)
    progress_text = st.empty()
    
    for index, row in base_df.iterrows():
        for column in selected_columns:
            base_value = str(row[column]) if pd.notnull(row[column]) else ''
            if base_value:
                if is_excel:
                    # Keep base_value as is, don't break into words
                    for compare_value, cell_location, original_value in compare_data:
                        compare_value = str(compare_value)
                        compare_words = re.findall(r'\w+', compare_value.lower())
                        
                        # Compare full base_value against each word and combinations
                        max_ratio = 0
                        best_match = ''
                        for i in range(len(compare_words)):
                            for j in range(i+1, len(compare_words)+1):
                                compare_phrase = ' '.join(compare_words[i:j])
                                ratio = fuzz.ratio(base_value.lower(), compare_phrase)
                                if ratio > max_ratio:
                                    max_ratio = ratio
                                    best_match = compare_phrase
                        
                        if max_ratio >= match_ratio:
                            results.append([column, base_value, f"{cell_location} ({best_match})", max_ratio])
                else:
                    # Existing logic for non-Excel files
                    text = compare_data.lower()
                    matches = []
                    for i in range(len(text) - len(base_value) + 1):
                        substring = text[i:i+len(base_value)]
                        ratio = fuzz.ratio(base_value.lower(), substring)
                        if ratio > match_ratio:
                            matches.append((ratio, substring, i))
                    
                    matches.sort(reverse=True, key=lambda x: x[0])
                    
                    if matches:
                        best_match = matches[0]
                        context_start = max(best_match[2] - 10, 0)
                        context_end = min(best_match[2] + len(best_match[1]) + 10, len(text))
                        context = text[context_start:context_end]
                        results.append([column, base_value, context, best_match[0]])
            
            # Update progress
            progress = (index * len(selected_columns) + selected_columns.index(column) + 1) / total_comparisons
            progress_bar.progress(progress)
            progress_text.text(f"Processing... {progress:.1%}")
    
    progress_bar.empty()
    progress_text.empty()
    return pd.DataFrame(results, columns=['Match Type', 'Base Value', 'File Context/Cell Location', 'Match Ratio'])

# Streamlit App
st.title("Travel Data Verifier")

# File Uploads
st.sidebar.header("Upload Files")
base_file = st.sidebar.file_uploader("Upload Base Excel File", type=['xlsx', 'xls'])
manifest_files = st.sidebar.file_uploader("Upload Manifest Files to Compare", type=['pdf', 'txt', 'htm', 'html', 'xlsx', 'xls', 'csv'], accept_multiple_files=True)

# Match Ratio Slider
match_ratio = st.sidebar.slider("Set Match Ratio (Recommended 80 or above for best results)", min_value=50, max_value=100, value=80, step=1)

if base_file:
    try:
        base_df = pd.read_excel(base_file, dtype=str)
    
        # Show column names with checkboxes
        st.sidebar.header("Select Columns to Compare")
        selected_columns = []
        for column in base_df.columns:
            if st.sidebar.checkbox(column, key=column):
                selected_columns.append(column)
        
        if manifest_files and selected_columns:
            all_results = pd.DataFrame()
            for manifest_file in manifest_files:
                file_extension = manifest_file.name.split('.')[-1].lower()

                # Extract text based on file type
                if file_extension == 'pdf':
                    compare_text = extract_text_from_pdf(manifest_file)
                    comparison_results = compare_data(base_df, compare_text, selected_columns, match_ratio)
                elif file_extension == 'txt':
                    compare_text = extract_text_from_txt(manifest_file)
                    comparison_results = compare_data(base_df, compare_text, selected_columns, match_ratio)
                elif file_extension in ['htm', 'html']:
                    compare_text = extract_text_from_html(manifest_file)
                    comparison_results = compare_data(base_df, compare_text, selected_columns, match_ratio)
                elif file_extension in ['xls', 'xlsx', 'csv']:
                    excel_data = extract_from_excel(manifest_file)
                    if excel_data == "encrypted":
                        password = st.sidebar.text_input("Enter password for encrypted Excel files", type="password")
                        if password:
                            excel_data = extract_from_excel(manifest_file, password)
                    if excel_data and excel_data != "encrypted":
                        comparison_results = compare_data(base_df, excel_data, selected_columns, match_ratio, is_excel=True)
                else:
                    st.error("Unsupported file type")
                    continue

                if not comparison_results.empty:
                    all_results = pd.concat([all_results, comparison_results], ignore_index=True)

            # Display Results
            if not all_results.empty:
                st.subheader("Comparison Results:")
                st.dataframe(all_results)
            
            else:
                st.info("No matches found.")
        elif not selected_columns:
            st.warning("Please select at least one column to compare.")
        elif not manifest_files:
            st.warning("Please upload manifest files to compare.")
    except Exception as e:
        st.error(f"An error occurred: {e}")